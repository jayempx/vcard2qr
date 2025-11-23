"""
vcard2qr GUI
===========================================
A PyQt5 application that lets you fill out contact information, generate a
vCard 3.0 payload, and render it as a customizable QR code.

Dependencies (install with pip):
    pip install PyQt5 qrcode[pil] pillow openpyxl

Author: ChatGPT (OpenAI o3)
Date: 2025‑08‑06
"""
#!/usr/bin/env python

import re
import sys
from pathlib import Path
from typing import Any, List, Mapping, Sequence, Tuple

from openpyxl import load_workbook
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QFormLayout,
    QHBoxLayout,
    QVBoxLayout,
    QLineEdit,
    QPushButton,
    QLabel,
    QColorDialog,
    QSpinBox,
    QScrollArea,
    QFileDialog,
    QMessageBox,
    QCheckBox,
    QFrame,
)

import qrcode
from qrcode.constants import ERROR_CORRECT_Q
from PIL import Image, ImageDraw

FIELD_HEADER_ALIASES = {
    "firstname": "First Name",
    "nome": "First Name",
    "lastname": "Last Name",
    "cognome": "Last Name",
    "organization": "Organization",
    "azienda": "Organization",
    "company": "Organization",
    "title": "Title",
    "email": "Email",
    "mobile": "Mobile",
    "cell": "Mobile",
    "cellphone": "Mobile",
    "cellulare": "Mobile",
    "switchboard": "Switchboard",
    "centralino": "Switchboard",
    "directoffice": "Direct Office",
    "directofficephone": "Direct Office",
    "office": "Direct Office",
    "address": "Address",
    "indirizzo": "Address",
    "linkedin": "LinkedIn",
}

TELEPHONE_TYPE_MAP = {
    "Mobile": "CELL",
    "Switchboard": "WORK",
    "Direct Office": "WORK;VOICE",
}

MINIMAL_VCARD = "BEGIN:VCARD\nVERSION:3.0\nEND:VCARD"


class VCardQRApp(QWidget):
    """Main application window."""

    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("vcard2qr")

        # ---------- Form fields ----------
        self.fields = {
            "First Name": QLineEdit(),
            "Last Name": QLineEdit(),
            "Organization": QLineEdit(),
            "Title": QLineEdit(),
            "Email": QLineEdit(),
            "Mobile": QLineEdit(),
            "Switchboard": QLineEdit(),
            "Direct Office": QLineEdit(),
            "Address": QLineEdit(),
            "LinkedIn": QLineEdit(),
        }
        self.custom_fields: List[Tuple[QLineEdit, QLineEdit]] = []

        # ---------- Layout scaffolding ----------
        root = QHBoxLayout(self)
        root.setSpacing(28)
        root.setContentsMargins(24, 24, 24, 24)

        # -- Left‑hand scrollable form
        form_container = QWidget()
        form_container.setObjectName("formContainer")
        self.form_layout = QFormLayout(form_container)
        for label, widget in self.fields.items():
            self.form_layout.addRow(label, widget)

        add_custom_btn = QPushButton("Add Custom Field")
        add_custom_btn.clicked.connect(self.add_custom_field)
        self.form_layout.addRow(add_custom_btn)

        scroll = QScrollArea()
        scroll.setWidget(form_container)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        root.addWidget(scroll, stretch=2)

        # -- Right‑hand controls & preview
        ctrl_col = QVBoxLayout()
        ctrl_col.setSpacing(14)
        ctrl_frame = QFrame()
        ctrl_frame.setObjectName("controlPanel")
        ctrl_frame.setLayout(ctrl_col)
        ctrl_frame.setFrameShape(QFrame.NoFrame)

        # Size selector
        ctrl_col.addWidget(QLabel("Size (px):"))
        self.size_spin = QSpinBox(minimum=128, maximum=2048, value=512) # type: ignore
        ctrl_col.addWidget(self.size_spin)

        # Corner radius
        ctrl_col.addWidget(QLabel("Corner radius (px):"))
        self.radius_spin = QSpinBox(minimum=0, maximum=20, value=4) # type: ignore
        ctrl_col.addWidget(self.radius_spin)

        # Colors
        self.fg_color = "#000000"
        self.bg_color = "#FFFFFF"
        self.transparent_bg = False

        fg_btn = QPushButton("Foreground color …")
        fg_btn.clicked.connect(self.pick_fg_color)
        ctrl_col.addWidget(fg_btn)

        self.bg_btn = QPushButton("Background color …")
        self.bg_btn.clicked.connect(self.pick_bg_color)
        ctrl_col.addWidget(self.bg_btn)

        self.transparent_checkbox = QCheckBox("No background (transparent)")
        self.transparent_checkbox.toggled.connect(self.toggle_transparent_background)
        ctrl_col.addWidget(self.transparent_checkbox)

        # Generate / actions
        gen_btn = QPushButton("Generate QR")
        gen_btn.clicked.connect(self.generate_qr)
        ctrl_col.addWidget(gen_btn)

        copy_btn = QPushButton("Copy to clipboard")
        copy_btn.clicked.connect(self.copy_to_clipboard)
        ctrl_col.addWidget(copy_btn)

        save_btn = QPushButton("Save as PNG …")
        save_btn.clicked.connect(self.save_png)
        ctrl_col.addWidget(save_btn)

        import_btn = QPushButton("Import from Excel …")
        import_btn.clicked.connect(self.import_from_excel)
        ctrl_col.addWidget(import_btn)

        ctrl_col.addStretch(1)

        # QR preview
        self.qr_label = QLabel(alignment=Qt.AlignCenter) # type: ignore
        self.qr_label.setMinimumSize(256, 256)
        ctrl_col.addWidget(self.qr_label, stretch=2)

        root.addWidget(ctrl_frame, stretch=1)

        self.current_qr: Image.Image | None = None
        self.apply_material_theme()

    # ---------------------------------------------------------------------
    # Utility helpers
    # ---------------------------------------------------------------------
    def add_custom_field(self) -> None:
        """Append a new key/value pair to the form."""
        key_edit, val_edit = QLineEdit(), QLineEdit()
        self.custom_fields.append((key_edit, val_edit))
        self.form_layout.addRow(key_edit, val_edit)

    def pick_fg_color(self) -> None:
        color = QColorDialog.getColor()
        if color.isValid():
            self.fg_color = color.name()

    def pick_bg_color(self) -> None:
        color = QColorDialog.getColor()
        if color.isValid():
            self.bg_color = color.name()

    def toggle_transparent_background(self, checked: bool) -> None:
        """Toggle whether the QR background should be transparent."""
        self.transparent_bg = checked
        self.bg_btn.setEnabled(not checked)

    def apply_material_theme(self) -> None:
        """Apply Material Design 3-inspired styling to the UI."""
        self.setStyleSheet(
            """
            QWidget {
                background-color: #f4f6fb;
                font-family: "Segoe UI", "Roboto", sans-serif;
                color: #1f2933;
            }
            #formContainer, #controlPanel {
                background-color: #ffffff;
                border-radius: 18px;
                border: 1px solid rgba(15, 23, 42, 0.1);
                padding: 18px;
            }
            #controlPanel {
                box-shadow: 0 16px 32px rgba(15, 23, 42, 0.12);
            }
            QScrollArea {
                border: none;
            }
            QLabel {
                font-size: 13px;
            }
            QLineEdit, QSpinBox {
                border-radius: 14px;
                border: 1px solid rgba(15, 23, 42, 0.2);
                padding: 6px 10px;
                background: #f7fafc;
            }
            QPushButton {
                border-radius: 16px;
                background-color: #1e88e5;
                color: white;
                padding: 12px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #1565c0;
            }
            QPushButton:disabled {
                background-color: rgba(30, 136, 229, 0.5);
            }
            QCheckBox {
                spacing: 8px;
            }
            """
        )

    # ---------------------------------------------------------------------
    # vCard & QR generation
    # ---------------------------------------------------------------------
    def build_vcard(self) -> str:
        """Compose a vCard 3.0 string from the form contents."""
        form_data = {
            label.replace("\u00a0", " ").strip(): widget.text().strip()
            for label, widget in self.fields.items()
        }
        custom_entries: List[Tuple[str, str]] = []
        for key_edit, val_edit in self.custom_fields:
            key = key_edit.text().strip()
            val = val_edit.text().strip()
            if key and val:
                custom_entries.append((key, val))
        return self._vcard_from_mapping(form_data, custom_entries)

    def _vcard_from_mapping(
        self,
        data: Mapping[str, str],
        custom_fields: Sequence[Tuple[str, str]],
    ) -> str:
        """Build a vCard payload from a mapping of standard and custom fields."""
        lines = ["BEGIN:VCARD", "VERSION:3.0"]

        fn = f"{data.get('First Name', '').strip()} {data.get('Last Name', '').strip()}".strip()
        if fn:
            lines.append(f"FN:{fn}")

        if org := data.get("Organization", "").strip():
            lines.append(f"ORG:{org}")
        if title := data.get("Title", "").strip():
            lines.append(f"TITLE:{title}")

        for key, vtype in TELEPHONE_TYPE_MAP.items():
            if num := data.get(key, "").strip():
                lines.append(f"TEL;TYPE={vtype}:{num}")

        if email := data.get("Email", "").strip():
            lines.append(f"EMAIL;TYPE=INTERNET:{email}")

        if addr := data.get("Address", "").strip():
            lines.append(f"ADR;TYPE=WORK:;;{addr};;;;")

        if url := data.get("LinkedIn", "").strip():
            lines.append(f"URL:{url}")

        for key, val in custom_fields:
            label = key.strip().upper().replace(" ", "_")
            value = val.strip()
            if label and value:
                lines.append(f"X-{label}:{value}")

        lines.append("END:VCARD")
        return "\n".join(lines)

    @staticmethod
    def _normalize_header(value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", value.strip().lower())

    @staticmethod
    def _stringify_cell(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _sanitize_filename(value: str) -> str:
        return re.sub(r"[^A-Za-z0-9_-]+", "_", value).strip("_")

    def import_from_excel(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import contacts from Excel",
            "",
            "Excel workbooks (*.xlsx)",
        )
        if not path:
            return

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            QMessageBox.critical(self, "Import failed", f"Unable to read {path}: {exc}")
            return

        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if len(rows) < 2:
            QMessageBox.warning(self, "No data", "The workbook must contain at least one contact row.")
            return

        header = rows[0]
        standard_columns: dict[int, str] = {}
        custom_columns: dict[int, str] = {}
        for idx, raw_header in enumerate(header):
            if raw_header is None:
                continue
            normalized = self._normalize_header(str(raw_header))
            if normalized and (canonical := FIELD_HEADER_ALIASES.get(normalized)):
                standard_columns[idx] = canonical
            else:
                cleaned = str(raw_header).strip()
                if cleaned:
                    custom_columns[idx] = cleaned

        if not standard_columns and not custom_columns:
            QMessageBox.warning(self, "Invalid header", "The spreadsheet header does not contain recognizable field names.")
            return

        out_dir = QFileDialog.getExistingDirectory(
            self,
            "Select output folder",
            str(Path(path).parent),
        )
        if not out_dir:
            return
        out_dir_path = Path(out_dir)

        saved = 0
        skipped = 0
        for row_idx, row in enumerate(rows[1:], start=2):
            row_data = {
                field: self._stringify_cell(row[idx]) if idx < len(row) else ""
                for idx, field in standard_columns.items()
            }
            row_custom: List[Tuple[str, str]] = []
            for idx, header_name in custom_columns.items():
                if idx < len(row):
                    value = self._stringify_cell(row[idx])
                    if value:
                        row_custom.append((header_name, value))

            vcard = self._vcard_from_mapping(row_data, row_custom)
            if vcard == MINIMAL_VCARD:
                skipped += 1
                continue

            qr = qrcode.QRCode(
                version=None,
                error_correction=ERROR_CORRECT_Q,
                box_size=10,
                border=4,
            )
            qr.add_data(vcard)
            qr.make(fit=True)

            img = self.render_rounded(
                qr,
                self.size_spin.value(),
                self.radius_spin.value(),
                self.transparent_bg,
            )
            self.current_qr = img

            base_value = f"{row_data.get('First Name', '')} {row_data.get('Last Name', '')}".strip()
            safe_base = self._sanitize_filename(base_value) or f"contact_{row_idx - 1}"
            target_path = out_dir_path / f"{safe_base}_{row_idx - 1}.png"
            try:
                img.save(target_path)
            except Exception as exc:
                QMessageBox.warning(
                    self,
                    "Save failed",
                    f"Could not write {target_path.name}: {exc}",
                )
                continue

            saved += 1

        if saved:
            summary = f"Saved {saved} QR code{'s' if saved != 1 else ''} to {out_dir_path}"
            if skipped:
                summary += f" (skipped {skipped} empty row{'s' if skipped != 1 else ''})"
            QMessageBox.information(self, "Import complete", summary)
        else:
            QMessageBox.warning(
                self,
                "No QR codes",
                "The imported rows did not contain enough data to build any QR codes.",
            )

    def generate_qr(self) -> None:
        vcard = self.build_vcard()
        if vcard == MINIMAL_VCARD:
            QMessageBox.warning(self, "Incomplete", "Add at least one field first.")
            return

        qr = qrcode.QRCode(
            version=None,
            error_correction=ERROR_CORRECT_Q,
            box_size=10,
            border=4,
        )
        qr.add_data(vcard)
        qr.make(fit=True)

        img = self.render_rounded(
            qr,
            self.size_spin.value(),
            self.radius_spin.value(),
            self.transparent_bg,
        )
        self.current_qr = img

        # Display in Qt label
        self.show_qr(img)

    def render_rounded(
        self,
        qr: qrcode.QRCode,
        size: int,
        radius: int,
        transparent_bg: bool,
    ) -> Image.Image:
        """Return a Pillow image with optional rounded modules."""
        matrix = qr.get_matrix()
        modules = len(matrix)
        box = size // (modules + 8)  # +8 for border 4 each side
        img_size = box * (modules + 8)
        mode = "RGBA" if transparent_bg else "RGB"
        background = (0, 0, 0, 0) if transparent_bg else self.bg_color
        img = Image.new(mode, (img_size, img_size), background)
        draw = ImageDraw.Draw(img)

        for r, row in enumerate(matrix):
            for c, val in enumerate(row):
                if val:
                    x0 = (c + 4) * box
                    y0 = (r + 4) * box
                    x1 = x0 + box
                    y1 = y0 + box
                    if radius:
                        draw.rounded_rectangle([x0, y0, x1, y1], radius=radius, fill=self.fg_color)
                    else:
                        draw.rectangle([x0, y0, x1, y1], fill=self.fg_color)

        if img_size != size:
            img = img.resize((size, size), Image.LANCZOS) # type: ignore
        return img

    # ---------------------------------------------------------------------
    # Clipboard / saving helpers
    # ---------------------------------------------------------------------
    def pil_to_qimage(self, img: Image.Image) -> QImage:
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGBA")
        data = img.tobytes("raw", img.mode)
        if img.mode == "RGBA":
            qimg = QImage(
                data,
                img.width,
                img.height,
                img.width * 4,
                QImage.Format_RGBA8888,
            )
        else:
            qimg = QImage(
                data,
                img.width,
                img.height,
                img.width * 3,
                QImage.Format_RGB888,
            )
        return qimg.copy()

    def show_qr(self, img: Image.Image) -> None:
        qimg = self.pil_to_qimage(img)
        pix = QPixmap.fromImage(qimg)
        self.qr_label.setPixmap(pix.scaled(self.qr_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)) # type: ignore

    def copy_to_clipboard(self) -> None:
        if self.current_qr is None:
            QMessageBox.information(self, "Nothing to copy", "Generate a QR first.")
            return
        qimg = self.pil_to_qimage(self.current_qr)
        QApplication.clipboard().setImage(qimg) # type: ignore
        QMessageBox.information(self, "Copied", "QR code copied to clipboard.")

    def save_png(self) -> None:
        if self.current_qr is None:
            QMessageBox.information(self, "Nothing to save", "Generate a QR first.")
            return
        path, _ = QFileDialog.getSaveFileName(self, "Save QR", "contact_qr.png", "PNG Images (*.png)")
        if path:
            self.current_qr.save(path)
            QMessageBox.information(self, "Saved", f"QR saved to {path}")


# -------------------------------------------------------------------------
# Entrypoint
# -------------------------------------------------------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = VCardQRApp()
    window.resize(900, 600)
    window.show()
    sys.exit(app.exec_())
